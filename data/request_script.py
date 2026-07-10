import requests
import pandas as pd
import json
import re
import time
import argparse
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from bs4 import BeautifulSoup

# ================= 配置区域 =================
# 1. 查询接口配置
QUERY_URL = "http://folcbla-asia.icbc:3012/"
QUERY_HEADERS = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "Accept-Language": "en-US,en;q=0.9",
    "Cache-Control": "max-age=0",
    "Connection": "keep-alive",
    "Content-Type": "application/x-www-form-urlencoded",
    "Origin": "null",
    "Upgrade-Insecure-Requests": "1",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36 Edg/138.0.0.0",
}
QUERY_CREDENTIALS = {
    "user": "cbla",
    "pass": "Oper1234",
    "UNIT": "PEP00110"
}

# 2. LLM 接口配置
LLM_URL = "http://123.192.49.9:8086/v1/chat/completions"
LLM_MODEL = "qwen3.5"

# 3. 文件路径配置
INPUT_FILE = "test.xlsx"
OUTPUT_FILE = "output_blacklist_result.xlsx"

QUERY_NAME_COLUMN = "English Name"  # 请根据实际 Excel 列名修改

# 4. 新增列定义
NEW_COLUMNS = [
    "If Hitted", "Hit#", "Exact Match", "Partial Match", "Not Match",
    "ID", "ORIGIN", "DESIGNATION", "PRIORITY", "CONFIDENTIALITY", 
    "CITY", "COUNTRY/REGION", "CATEGORIES", "KEYWORDS", "TYPE", 
    "ADDRESS", "SEARCHED CODES", "BIC CODES", "NATIONAL ID", 
    "PASSPORT NO", "PLACE OF BIRTH", "DATE OF BIRTH", 
    "USER INFO1", "USER INFO2", "OFFICIAL REFERENCE", "ADDITIONAL INFO"
]

# ================= 核心逻辑函数 =================

def query_blacklist_api(name: str):
    """调用内部黑名单查询接口"""
    data = {
        "name": name,
        "ADDR": "", "city": "", "STAT": "", "CTRY": "", "CODE": "",
        "BIC": "", "NID": "", "PSP": "", "TYPE": "", "DATEOFBIRTH": "",
        "UNIT": QUERY_CREDENTIALS["UNIT"],
        "user": QUERY_CREDENTIALS["user"],
        "pass": QUERY_CREDENTIALS["pass"],
    }
    try:
        response = requests.post(QUERY_URL, headers=QUERY_HEADERS, data=data, timeout=30)
        if response.status_code == 200:
            return response.text
        else:
            print(f"  [查询失败] Status {response.status_code}: {name}")
            return None
    except Exception as e:
        print(f"  [网络异常] {e}")
        return None


def clean_html_with_bs4(html_content: str) -> str:
    """
    使用 BeautifulSoup 清理 HTML，提取核心数据内容
    移除冗余的 Word/FrontPage 样式标签，保留文本和表格结构
    """
    try:
        soup = BeautifulSoup(html_content, 'html.parser')
        
        # 1. 移除脚本和样式标签
        for tag_name in ['script', 'style', 'meta', 'link']:
            for tag in soup.find_all(tag_name):
                tag.decompose()
        
        # 2. 清理所有标签中的冗余属性
        attrs_to_remove_patterns = [
            'style', 'width', 'height', 'bgcolor', 'border', 
            'cellpadding', 'cellspacing', 'xmlns', 'fpstyle',
        ]
        attrs_to_remove_regex = [
            re.compile(r'^mso-'), re.compile(r'^v:'), 
            re.compile(r'^o:'), re.compile(r'^w:'),
            re.compile(r'^u[0-9]+:')
        ]
        
        for tag in soup.find_all():  # 不使用 True 参数
            attrs_to_del = []
            for attr in tag.attrs:
                # 检查是否在移除列表中
                if attr in attrs_to_remove_patterns:
                    attrs_to_del.append(attr)
                # 检查是否匹配正则模式
                elif any(pattern.match(attr) for pattern in attrs_to_remove_regex):
                    attrs_to_del.append(attr)
            for attr in attrs_to_del:
                del tag[attr]
        
        # 3. 清理文本中的冗余标签（如 <o:p>, <u1:p>, <u2:p>）
        o_p_pattern = re.compile(r'^o:p$')
        u_p_pattern = re.compile(r'^u[0-9]+:p$')
        
        for tag in soup.find_all():
            tag_name = tag.name if tag.name else ''
            if o_p_pattern.match(tag_name) or u_p_pattern.match(tag_name):
                tag.replace_with('')
        
        # 4. 清理空标签（只保留有内容的标签）
        tags_to_decompose = []
        for tag in soup.find_all():
            # 使用 tag.contents 获取子元素列表（更兼容）
            if tag.name and not tag.get_text(strip=True) and len(tag.contents) == 0:
                tags_to_decompose.append(tag)
        for tag in tags_to_decompose:
            tag.decompose()
        
        # 5. 获取清理后的 HTML
        cleaned_html = str(soup)
        
        # 6. 进一步清理：移除连续空白和多余换行
        cleaned_html = re.sub(r'\n\s*\n\s*\n+', '\n\n', cleaned_html)
        cleaned_html = re.sub(r'[ \t]+', ' ', cleaned_html)
        
        return cleaned_html.strip()
        
    except Exception as e:
        import traceback
        print(f"  [HTML 清理异常] {e}")
        traceback.print_exc()
        return html_content


def extract_fields_with_llm(html_content: str):
    """
    调用大模型提取字段。
    关键修改：要求返回一个 HIT 列表 (List of Hits)，以支持多结果场景。
    """
    # 先使用 BeautifulSoup 清理 HTML
    cleaned_html = clean_html_with_bs4(html_content)
    
    # 检查标题判断是否命中
    if "No hit found" in cleaned_html or "No hit found" in html_content:
        print("  [无命中] No hit found in response")
        return []
    
    # 构建 prompt - 使用字符串拼接避免 f-string 三重引号问题
    prompt = """# Role
你是一个专业的金融合规数据提取助手。从 HTML 中提取**所有**黑名单命中结果（HITS）。

# Input HTML (已清理)
"""
    prompt += cleaned_html
    prompt += """
# Rules
1. **输出结构**：必须返回一个 JSON 列表 `[]`。列表中的每个元素代表一个 HIT 结果。
   如果未找到任何命中，返回空列表 `[]`。
   
2. **每个 HIT 对象包含以下字段**：
   - ID, ORIGIN, DESIGNATION, PRIORITY, CONFIDENTIALITY, CITY, COUNTRY/REGION, CATEGORIES, KEYWORDS, TYPE, ADDRESS, SEARCHED CODES, BIC CODES, NATIONAL ID, PASSPORT NO, PLACE OF BIRTH, DATE OF BIRTH, USER INFO1, USER INFO2, OFFICIAL REFERENCE, ADDITIONAL INFO
   - NAME_LIST: 该 HIT 下所有出现的名字变体（数组格式）。

3. **字符连续性约束（至关重要）**：
   - **严禁**在中英文、中文字符与数字之间添加任何多余的空格。
   - 保持原始数据的紧凑性。例如："GMT 智能系统 AI.mp4"。
   - 地址、姓名等长文本，去除字母与数字/汉字间的空格。

4. **输出格式**：
   - 仅输出纯 JSON 列表，无 Markdown 标记（如 ```json），无解释性文字。
   - 示例：
   [
     {
       "ID": "123",
       "NAME_LIST": ["JohnDoe", "DoeJohn"],
       "ADDRESS": "No.1Road,Beijing",
       ...
     },
     {
       "ID": "456",
       "NAME_LIST": ["JaneSmith"],
       ...
     }
   ]
"""

    payload = {
        "model": LLM_MODEL,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.0
    }

    try:
        resp = requests.post(LLM_URL, json=payload, headers={"Content-Type": "application/json"}, timeout=600)
        resp.raise_for_status()
        
        # 调试：打印响应内容
        raw_content = resp.text
        if len(raw_content) < 100:
            print(f"  [LLM 响应异常短] {raw_content[:200]}")
        
        try:
            resp_data = resp.json()
            content = resp_data['choices'][0]['message']['content']
        except (KeyError, json.JSONDecodeError) as e:
            print(f"  [LLM 响应格式错误] {e}")
            print(f"  原始响应：{raw_content[:500]}")
            return []
        
        # 清理可能存在的 Markdown 标记
        content = re.sub(r'^```json\s*', '', content.strip())
        content = re.sub(r'\s*```$', '', content.strip())
        content = re.sub(r'^```', '', content.strip())
        content = re.sub(r'```$', '', content.strip())
        
        # 调试：打印清理后的内容
        if len(content) < 50:
            print(f"  [LLM 内容异常短] {content[:200]}")
        
        data = json.loads(content)
        
        # 确保返回的是列表
        if isinstance(data, list):
            return data
        else:
            # 如果模型返回了单个对象而不是列表，尝试包装成列表
            if isinstance(data, dict) and "error" not in data:
                return [data]
            return []
            
    except json.JSONDecodeError as e:
        print(f"  [JSON 解析错误] {e}")
        print(f"  尝试解析的内容：{content[:300] if 'content' in dir() else 'N/A'}")
        return []
    except requests.exceptions.RequestException as e:
        print(f"  [LLM 请求错误] {e}")
        return []
    except Exception as e:
        print(f"  [LLM 解析错误] {e}")
        import traceback
        traceback.print_exc()
        return []


def normalize_name(name: str):
    """标准化名字：转小写，去标点，去多余空格，转为单词集合"""
    if not name:
        return set()
    name = name.lower()
    # 去除非字母数字字符（保留空格用于分词）
    name = re.sub(r'[^\w\s]', '', name)
    # 分割并去重
    return set(name.split())


def classify_match(query_name: str, hit_names_list: list):
    """
    分类匹配逻辑
    返回：'Exact Match', 'Partial Match', 'Not Match'
    """
    query_words = normalize_name(query_name)
    if not query_words:
        return "Not Match"

    best_type = "Not Match"

    for hit_name in hit_names_list:
        hit_words = normalize_name(hit_name)
        if not hit_words:
            continue
        
        # Exact: 查询的所有单词都在结果中
        if query_words.issubset(hit_words):
            return "Exact Match"
        
        # Partial: 有交集
        if query_words.intersection(hit_words):
            if best_type == "Not Match":
                best_type = "Partial Match"
    
    return best_type


def process_excel(input_file: str, output_file: str, max_rows: int = None):
    """主处理流程
    
    Args:
        input_file: 源 Excel 文件路径
        output_file: 目标 Excel 文件路径
        max_rows: 最大处理条目数量，None 表示处理全部
    """
    print(f"正在读取文件：{input_file}")
    try:
        df = pd.read_excel(input_file)
    except FileNotFoundError:
        print(f"错误：找不到文件 {INPUT_FILE}")
        return

    if QUERY_NAME_COLUMN not in df.columns:
        print(f"错误：Excel 中未找到列 '{QUERY_NAME_COLUMN}'")
        return

    # 如果指定了处理条目数量，则限制行数
    if max_rows is not None and max_rows > 0:
        df = df.head(max_rows)
        print(f"限制处理条目数量为：{max_rows}")
    
    total_rows = len(df)
    results = []
    
    # 遍历每一行
    for index, row in df.iterrows():
        query_name = str(row[QUERY_NAME_COLUMN]).strip()
        if not query_name:
            continue

        print(f"处理 [{index+1}/{total_rows}]: {query_name}")

        # 1. 查询接口
        html = query_blacklist_api(query_name)
        if not html:
            # 查询失败
            new_row = row.to_dict()
            new_row["If Hitted"] = "ERROR: QUERY_FAILED"
            for col in NEW_COLUMNS[1:]:
                new_row[col] = ""
            results.append(new_row)
            continue

        # 2. LLM 解析 (返回 HIT 列表)
        hits_list = extract_fields_with_llm(html)
        
        if not hits_list:
            # 未命中或解析为空
            new_row = row.to_dict()
            new_row["If Hitted"] = "NO HIT DETECTED"
            for col in NEW_COLUMNS[1:]:
                new_row[col] = ""
            results.append(new_row)
            continue

        # 3. 命中处理：遍历每个 HIT
        # 注意：这里 hits_list 是一个列表，包含多个 HIT 对象
        for hit_idx, hit_data in enumerate(hits_list):
            hit_num = hit_idx + 1
            
            # 获取名字列表
            name_list = hit_data.get("NAME_LIST", [])
            if not name_list:
                # 有数据但没名字，视为不匹配或跳过，这里按不匹配处理
                match_type = "Not Match"
                names_str = ""
            else:
                # 匹配逻辑
                match_type = classify_match(query_name, name_list)
                # 将名字列表合并为字符串，换行显示
                names_str = "\n".join(name_list)

            # 准备新行数据 (复制原行)
            new_row = row.to_dict()
            new_row["If Hitted"] = "POTENTIAL HITS FOUND"
            new_row["Hit#"] = hit_num
            
            # 填充匹配列
            new_row["Exact Match"] = ""
            new_row["Partial Match"] = ""
            new_row["Not Match"] = ""
            
            if match_type == "Exact Match":
                new_row["Exact Match"] = names_str
            elif match_type == "Partial Match":
                new_row["Partial Match"] = names_str
            else:
                new_row["Not Match"] = names_str

            # 填充其他详情列
            detail_cols = ["ID", "ORIGIN", "DESIGNATION", "PRIORITY", "CONFIDENTIALITY", 
                           "CITY", "COUNTRY/REGION", "CATEGORIES", "KEYWORDS", "TYPE", 
                           "ADDRESS", "SEARCHED CODES", "BIC CODES", "NATIONAL ID", 
                           "PASSPORT NO", "PLACE OF BIRTH", "DATE OF BIRTH", 
                           "USER INFO1", "USER INFO2", "OFFICIAL REFERENCE", "ADDITIONAL INFO"]
            
            for col in detail_cols:
                val = hit_data.get(col, "")
                if isinstance(val, list):
                    val = "\n".join(str(v) for v in val)
                new_row[col] = str(val) if val else ""
            print(new_row)
            results.append(new_row)

        # 简单延时，避免接口限流
        time.sleep(0.5)

    # 4. 构建最终 DataFrame
    final_df = pd.DataFrame(results)

    # 确保列顺序正确
    original_cols = [c for c in df.columns if c not in NEW_COLUMNS]
    final_cols = original_cols + NEW_COLUMNS
    final_df = final_df[final_cols]

    # 5. 保存 Excel 并设置格式
    print(f"正在保存结果到：{output_file}")
    final_df.to_excel(output_file, index=False)

    # 使用 openpyxl 处理行高和换行
    wb = load_workbook(output_file)
    ws = wb.active

    # 设置列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50)

    # 设置所有新增列的单元格格式：换行显示，不自动调整行高
    start_col_idx = len(original_cols) + 1
    end_col_idx = len(final_cols)

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(start_col_idx, end_col_idx + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell.alignment = Alignment(wrap_text=True)
                # 不设置 row_height，保持默认，实现"双击显示"效果

    wb.save(output_file)
    print("处理完成！")


if __name__ == "__main__":
    # 解析命令行参数
    parser = argparse.ArgumentParser(description="黑名单查询脚本")
    parser.add_argument("-i", "--input", required=True, help="源 Excel 文件路径")
    parser.add_argument("-o", "--output", required=True, help="目标 Excel 文件路径")
    parser.add_argument("-n", "--num", type=int, default=None, help="处理条目数量（不输入则处理整个文件）")
    
    args = parser.parse_args()
    
    process_excel(args.input, args.output, args.num)